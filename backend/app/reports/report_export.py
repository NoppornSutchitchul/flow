"""Generate table-based PDF / Excel exports (not browser print)."""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_FONT_PATH = Path(__file__).resolve().parent / "assets" / "fonts" / "Sarabun-Regular.ttf"
_FONT_NAME = "Sarabun"
_font_registered = False


class ExportSectionModel(BaseModel):
    heading: str | None = None
    columns: list[str] = Field(default_factory=list)
    rows: list[list[str | int | float]] = Field(default_factory=list)


class ExportDocumentModel(BaseModel):
    title: str
    subtitle: str | None = None
    meta: list[str] = Field(default_factory=list)
    sections: list[ExportSectionModel] = Field(default_factory=list)


def _register_font() -> str:
    global _font_registered
    if _font_registered:
        return _FONT_NAME
    if _FONT_PATH.is_file():
        pdfmetrics.registerFont(TTFont(_FONT_NAME, str(_FONT_PATH)))
        _font_registered = True
        return _FONT_NAME
    return "Helvetica"


def _safe_sheet_name(name: str, index: int) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "", name or f"Sheet{index + 1}")[:31]
    return cleaned or f"Sheet{index + 1}"


def build_xlsx_bytes(doc: ExportDocumentModel) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    if not doc.sections:
        ws = wb.create_sheet("Report")
        ws.append([doc.title])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for i, section in enumerate(doc.sections):
        ws = wb.create_sheet(_safe_sheet_name(section.heading or "", i))
        row_idx = 1
        if i == 0:
            ws.cell(row=row_idx, column=1, value=doc.title).font = Font(bold=True, size=14)
            row_idx += 1
            if doc.subtitle:
                ws.cell(row=row_idx, column=1, value=doc.subtitle)
                row_idx += 1
            for line in doc.meta:
                ws.cell(row=row_idx, column=1, value=line)
                row_idx += 1
            row_idx += 1

        if section.heading and i > 0:
            ws.cell(row=row_idx, column=1, value=section.heading).font = Font(bold=True)
            row_idx += 2

        if section.columns:
            for col_i, col in enumerate(section.columns, 1):
                cell = ws.cell(row=row_idx, column=col_i, value=col)
                cell.font = Font(bold=True)
            row_idx += 1

        for row in section.rows:
            for col_i, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_i, value=val)
            row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(doc: ExportDocumentModel) -> bytes:
    font = _register_font()
    buf = io.BytesIO()
    pdf = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    title_style = ParagraphStyle(
        "Title",
        fontName=font,
        fontSize=14,
        leading=18,
        spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        "Sub",
        fontName=font,
        fontSize=10,
        leading=13,
        textColor=colors.grey,
        spaceAfter=4,
    )
    head_style = ParagraphStyle(
        "SectionHead",
        fontName=font,
        fontSize=11,
        leading=14,
        spaceBefore=8,
        spaceAfter=4,
    )
    story: list[Any] = []
    story.append(Paragraph(doc.title.replace("&", "&amp;"), title_style))
    if doc.subtitle:
        story.append(Paragraph(doc.subtitle.replace("&", "&amp;"), sub_style))
    for line in doc.meta:
        story.append(Paragraph(line.replace("&", "&amp;"), sub_style))

    for section in doc.sections:
        if section.heading:
            story.append(Paragraph(section.heading.replace("&", "&amp;"), head_style))
        if not section.columns:
            continue
        table_data: list[list[Any]] = [section.columns]
        for row in section.rows:
            table_data.append([str(c) for c in row])
        col_count = len(section.columns)
        width = (A4[0] - 28 * mm) / max(col_count, 1)
        tbl = Table(
            table_data,
            colWidths=[width] * col_count,
            repeatRows=1,
        )
        tbl.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, 0), font, 9),
                    ("FONT", (0, 1), (-1, -1), font, 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d2d2d")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f5f0")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ],
            ),
        )
        story.append(tbl)
        story.append(Spacer(1, 6 * mm))

    pdf.build(story)
    return buf.getvalue()
